import sys
import smtplib
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from os.path import basename
import openpyxl as excel

ROW_PAYLOAD_START = 2
COL_MAIL_TO = 1
COL_MAIL_SUBJECT = 2
COL_MAIL_BODY = 3
COL_MAIL_ATTACHMENT = 4

def load_excel(file_path):
    try:
        wb = excel.load_workbook(file_path, data_only=True)
        ws_setting = wb['設定シート']
        ws_main = wb['メール本文シート']
        return ws_setting, ws_main
    except Exception as e:
        sys.exit(f"Excelファイルの読み込み中にエラーが発生しました: {e}")

def get_smtp_connection(ws_setting):
    try:
        mail_from = ws_setting['b1'].value
        mail_server = ws_setting['b2'].value
        port = ws_setting['b3'].value
        mail_id = ws_setting['b4'].value
        mail_pass = ws_setting['b5'].value
        
        server = smtplib.SMTP(mail_server, port)
        server.starttls()  # 必要に応じてTLSを有効にします
        server.login(mail_id, mail_pass)
        return server, mail_from
    except Exception as e:
        sys.exit(f"SMTPサーバー接続中にエラーが発生しました: {e}")

def send_mail(server, mail_from, mail_to, subject, body, attachment=None):
    try:
        if attachment:
            message = MIMEMultipart()
            message.attach(MIMEText(body))
            
            with open(attachment, 'rb') as f:
                part = MIMEApplication(f.read(), Name=basename(attachment))
                part['Content-Disposition'] = f'attachment; filename="{basename(attachment)}"'
                message.attach(part)
        else:
            message = MIMEText(body)
        
        message['Subject'] = subject
        message['From'] = mail_from
        message['To'] = mail_to
        
        server.sendmail(mail_from, mail_to, message.as_string())
    except Exception as e:
        print(f"{mail_to}へのメール送信中にエラーが発生しました: {e}")

def main():
    if len(sys.argv) != 2:
        sys.exit("使用方法: python script.py <Excelファイルパス>")
    
    file_path = sys.argv[1]
    ws_setting, ws_main = load_excel(file_path)
    server, mail_from = get_smtp_connection(ws_setting)

    try:
        for row in range(ROW_PAYLOAD_START, ws_main.max_row + 1):
            mail_to = ws_main.cell(row=row, column=COL_MAIL_TO).value
            subject = ws_main.cell(row=row, column=COL_MAIL_SUBJECT).value
            body = ws_main.cell(row=row, column=COL_MAIL_BODY).value
            attachment = ws_main.cell(row=row, column=COL_MAIL_ATTACHMENT).value
            
            if mail_to:
                send_mail(server, mail_from, mail_to, subject, body, attachment)
    except Exception as e:
        print(f"メール送信処理中にエラーが発生しました: {e}")
    finally:
        server.quit()

if __name__ == "__main__":
    main()
